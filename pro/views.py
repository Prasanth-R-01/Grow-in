# from msilib.schema import Media
from django.shortcuts import render
from django.contrib.auth.models import User ,auth
from django.shortcuts import render,redirect
from django.contrib import messages
import requests
import joblib
import pandas as pd
import numpy as np
from .forms import *
import csv
from .models import *
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
import os
from werkzeug.utils import secure_filename
from django.core.files.images import ImageFile
# from msilib.schema import Media
from keras.models import load_model
from keras.applications.imagenet_utils import preprocess_input, decode_predictions
from keras.models import load_model
from keras.preprocessing import image
from keras import utils
import os,glob
import openpyxl
# from sklearn import metrics
# import seaborn as sns
# import matplotlib.pyplot as plt
# Create your views here.


def cropInfo(request):
    context = {}
    if request.method == 'POST':
        name = request.POST['name']

        if cropDetails.objects.filter(name=name).exists():
            crop = cropDetails.objects.get(name=name)
            context['name'] = crop.name
            context['alternateName'] = crop.alternateName
            context['sowInstructions'] = crop.sowInstructions
            context['spaceInstructions'] = crop.spaceInstructions
            context['harvestInstructions'] = crop.harvestInstructions
            context['compatiblePlants'] = crop.compatiblePlants
            context['avoidInstructions'] = crop.avoidInstructions
            context['culinaryHints'] = crop.culinaryHints
            context['culinaryPreservation'] = crop.culinaryPreservation
            context['url'] = crop.url

            return render(request, 'amaranth.html' , context)
            
    return render(request, 'amaranth.html' , context)

def model_predict(img_path, model):
    # img = image.load_img(img_path, target_size=(224, 224))

    img = utils.load_img(img_path, target_size=(256, 256))

    # Preprocessing the image
    x = utils.img_to_array(img)
    # x = np.true_divide(x, 255)
    x = np.expand_dims(x, axis=0)

    # Be careful how your trained model deals with the input
    # otherwise, it won't make correct prediction!
    x = preprocess_input(x, mode='caffe')

    preds = model.predict(x)
    return preds

def dp_Img(request):
    d = {}
    if request.method == "POST":
        f = request.FILES['dpimg']
        file_name = f.name
        print(dpDetails.objects.filter(name = file_name).exists())
        if not dpDetails.objects.filter(name = file_name).exists():
            print("In If")
            obj = dpDetails.objects.create(
                name = file_name , 
                dpimg  = f,
            )
            obj.save();
        
        dp = dpDetails.objects.get(name = file_name)
        dpimg =  dp.dpimg
        print(type(dpimg))
        print(dpimg)
        # Model saved with Keras model.save()
        MODEL_PATH = 'Templates/kitchen_garden_vgg19.h5'

        # Load your trained model
        model = load_model(MODEL_PATH)
        model.make_predict_function()          # Necessary
        print('Model loaded. Start serving...')
        print(model.summary())

        # Save the file to ./uploads
        # basepath = os.path.dirname('images')
        # file_path = os.path.join(
        #     basepath, 'images', secure_filename(file_name))
        
        # img =  ImageFile(f)
        
        # print(file_path)
        # Make prediction
        file_name = str(dpimg)
        file_path = "./data/" + file_name #"D:/AGRI/AGRI/agri/data/"+file_name
        print(file_path)
        print("File Name", file_name)
        
        
        preds = model_predict(file_path, model)

        # Process your result for human
        pred_class = preds.argmax(axis=-1)            # Simple argmax
        # pred_class = decode_predictions(preds, top=1)   # ImageNet Decode

        # result = str(pred_class[0][0][1])               # Convert to string
        print(pred_class)
        ref = {   0: 'Bacterial_Spot',
            1: 'Black Rot',
            2: 'Early Blight',
            3: 'Healthy',
            4: 'Late Blight',
            5: 'Leaf Mold',
            6: 'Leaf Spots',
            7: 'Mosaic Virus',
            8: 'Powdery Mildew',
            9: 'Rusts'
        }
        result = ref[pred_class[0]]
        print(ref[pred_class[0]])

        dp.delete()
        
        directory = "./data/data/"
        filelist = glob.glob(os.path.join(directory, "*"))
        for f in filelist:
            os.remove(f)
            
        print('result : ' +  result)
        obj = diseaseInfo.objects.get(diseaseName=result)
        name =obj.diseaseName
        cont = obj.context
        d = {'name':name , 'context' : cont}

    return render(request,'disease_pred.html',d)
    #     name = 'test'
    #     print(dpimg)
    #     obj =  dpDetails.objects.create(
    #         name=name,
    #         dpimg = dpimg
    #     )    
    #     obj.save();

    # return render(request,'disease_pred.html' )

def dp(request):
    return render(request,'disease_pred.html')


def crop_rec(request):
    if request.method == 'POST':
        lat = request.POST['lat']
        lon = request.POST['lon']
        area = float(request.POST['area'])
        unit = request.POST['unit']
        quantity = float(request.POST['quantity'])
        quality = request.POST['quality']
        soil = request.POST['soil']
        soiltestresult = request.POST['soil-test-result']
        if soiltestresult == 'Yes':
            ph = float(request.POST['ph'])
        else:
            ph = 0

        temp_min_sum = 0
        temp_max_sum = 0
        humidity_sum = 0
        rainfall_sum = 0

        try:
            url_w = "https://pro.openweathermap.org/data/2.5/forecast/climate?lat="+str(lat)+"&lon="+str(lon)+"&appid=598e3f9cf0c924b6ddb56c192425f2f1&units=metric"

            import urllib.request, json 
            with urllib.request.urlopen(url_w) as link:
                data = json.loads(link.read().decode())

            l=len(data["list"])
            
            
            for i in range(0,l):
                temp_min_sum += data["list"][i]["temp"]["min"]    
                temp_max_sum += data["list"][i]["temp"]["max"]    
                humidity_sum += data["list"][i]["humidity"]

                if "rain" in data["list"][i]:
                    rainfall_sum += data["list"][i]["rain"]

        except:   
            url_w = "https://api.openweathermap.org/data/2.5/onecall?lat="+str(lat)+"&lon="+str(lon)+"&exclude=current,minutely,hourly,alerts&units=metric&appid=598e3f9cf0c924b6ddb56c192425f2f1#"

            import urllib.request, json 
            with urllib.request.urlopen(url_w) as link:
                data = json.loads(link.read().decode())

            l=len(data["daily"])

            for i in range(0,l):
                temp_min_sum += data["daily"][i]["temp"]["min"]
                temp_max_sum += data["daily"][i]["temp"]["max"]
                humidity_sum += data["daily"][i]["humidity"]

                if "rain" in data["daily"][i]:
                    rainfall_sum += data["daily"][i]["rain"]


        temp_min = round(temp_min_sum/l,2)
        temp_max = round(temp_max_sum/l,2)
        humidity = round(humidity_sum/l,2)

        weatherbit_flag = 0
        ambee_flag = 0
        is_soil_temp_available = 0
        soil_temp = 0

        #weatherbit
        #def weatherbit():
        try:
            from datetime import date, timedelta
            import urllib.request, json

            dte = date.today() - timedelta(1)   #current_date
            dts = dte - timedelta(7)            #7_days_before_current_date 

            # api_key_cpy = "ce314542553946b485c3b5d6ae1524b6"
            api_key = "84fba09db7e74f5da8b1c1dc47e9eccc"

            url_s = "https://api.weatherbit.io/v2.0/history/agweather?lat="+str(lat)+"&lon="+str(lon)+"&start_date="+str(dts)+"&end_date="+str(dte)+"&key="+str(api_key)

            with urllib.request.urlopen(url_s) as link:
                details = json.loads(link.read().decode())

            s = details["data"][0]["soilt_0_10cm"] + details["data"][1]["soilt_0_10cm"] + details["data"][2]["soilt_0_10cm"] + details["data"][3]["soilt_0_10cm"] + details["data"][4]["soilt_0_10cm"] + details["data"][5]["soilt_0_10cm"] + details["data"][6]["soilt_0_10cm"]
            soil_temp = round(s/7,2)
            weatherbit_flag=0
        except:
            weatherbit_flag=1

        # weatherbit()
        print(weatherbit_flag)
        if weatherbit_flag == 1:
            
            #ambee
            #def ambee():
            try:

                url = "https://api.ambeedata.com/soil/latest/by-lat-lng"
                querystring = {"lat":lat,"lng":lon}

                headers = {
                    'x-api-key': "240907f36f6cddb07b038e30a276467fdb5ec1e1a9b3a8e8f8b7beecceda062c",
                    'Content-type': "application/json"
                    }
                response = requests.request("GET", url, headers=headers, params=querystring)
                # print(response.text)

                import ast
                a=ast.literal_eval(response.text)
                soil_temp=a['data'][0]['soil_temperature']
                ambee_flag=0

            except:
                ambee_flag=1

            #ambee()    
            if ambee_flag == 1:
                is_soil_temp_available = 1
        
        soil_temp_avg =  2.508895
        if is_soil_temp_available == 0:
            min_soil_temp = soil_temp-round(soil_temp_avg,1)
            max_soil_temp = soil_temp+round(soil_temp_avg,1)
        # print("Min Soil temp: {}\nMax Soil temp: {}".format(min_soil_temp, max_soil_temp))

        c = int(0)
        water_avail_high = int(0)
        water_avail_medium = int(0)
        water_avail_low = int(0)

        if unit=="cent":
            c=area
        elif unit=="Sq feet":
            c=(area*0.0023)
        else:
            c=(area*0.02471)

        if quantity>=c*40:
            print("HIGH AVAIL.")
            water_avail_high=1
            water_avail_medium=0
            water_avail_low=0
            
        elif quantity>c*25 and quantity<c*40:
            print("MED AVAIL.")
            water_avail_high=1
            water_avail_medium=1
            water_avail_low=0
            
        else:
            print("LOW AVAIL.")
            water_avail_high=1
            water_avail_medium=1
            water_avail_low=1
        
        

        if quality == 'drinkable':
            water_qual_drinkable = 1
            water_qual_less_salinity = 0
            water_qual_more_salinity = 0
        elif quality == 'lessSalinity':
            water_qual_drinkable = 0
            water_qual_less_salinity = 1
            water_qual_more_salinity = 0
        elif quality == 'moreSalinity':
            messages.info( request , "Water is more saline, so no crop can be grown")
            return render(request, 'crop_recom.html')
            

        # print("Drinkable: {}, Less Saline: {}, More Saline: {}".format(water_qual_drinkable, water_qual_less_salinity, water_qual_more_salinity))
    
        if soil == 'sandy':
            sandy_soil = 1
            loamy_soil = 0
            clay_soil = 0
        elif soil == 'loamy':
            sandy_soil = 0
            loamy_soil = 1
            clay_soil = 0
        elif soil == 'clay':
            sandy_soil = 0
            loamy_soil = 0
            clay_soil = 1
        
        ph_avg = 0.24
        if soiltestresult == 'Yes':
            min_ph = ph - round(ph_avg,1)
            max_ph = ph + round(ph_avg,1)
        # print("Min Ph: {}\nMax Ph: {}".format(min_ph, round(max_ph,2)))

        humidity_avg = 4.07
        min_humidity = humidity - round(humidity_avg,1)
        max_humidity = humidity + round(humidity_avg,1)
        # print("Min Humidity: {}\nMax Humidity: {}".format(min_humidity, round(max_humidity,2)))


        if soiltestresult == 'Yes' and is_soil_temp_available == 0 : #svm_with_soil_test_with_soil_temp
            inputs=[min_ph, max_ph, temp_min, temp_max, min_soil_temp, max_soil_temp, min_humidity, max_humidity, water_avail_high, 
                        water_avail_medium, water_avail_low, water_qual_drinkable, water_qual_less_salinity, 
                        water_qual_more_salinity, sandy_soil, loamy_soil, clay_soil]
            model = joblib.load('Templates/svm_with_soil_test_with_soil_temp.pkl')

        elif soiltestresult == 'No' and is_soil_temp_available == 0:
            inputs=[temp_min, temp_max, min_soil_temp, max_soil_temp, min_humidity, max_humidity, water_avail_high, 
                        water_avail_medium, water_avail_low, water_qual_drinkable, water_qual_less_salinity, 
                        water_qual_more_salinity, sandy_soil, loamy_soil, clay_soil]
            model = joblib.load('Templates/svm_without_soil_test_with_soil_temp.pkl')

        elif soiltestresult == 'Yes' and is_soil_temp_available == 1:
            inputs=[min_ph, max_ph, temp_min, temp_max, min_humidity, max_humidity, water_avail_high, 
                        water_avail_medium, water_avail_low, water_qual_drinkable, water_qual_less_salinity, 
                        water_qual_more_salinity, sandy_soil, loamy_soil, clay_soil] 
            model = joblib.load('Templates/svm_with_soil_test_without_soil_temp.pkl')

        elif soiltestresult == 'No' and is_soil_temp_available == 1: 
            inputs=[temp_min, temp_max, min_humidity, max_humidity, water_avail_high, 
                        water_avail_medium, water_avail_low, water_qual_drinkable, water_qual_less_salinity, 
                        water_qual_more_salinity, sandy_soil, loamy_soil, clay_soil] 
            model = joblib.load('Templates/svm_without_soil_test_without_soil_temp.pkl')        
           
        r = model.predict_proba(np.array([inputs]))
        pred_table_svm = pd.DataFrame()
        pred_table_svm['crop'] = model.classes_
        pred_table_svm['prob'] = r[0]*100
        pred_table_svm.sort_values(by='prob', ascending=False, inplace=True)
        res = pred_table_svm[:10]
        print(res)
        res = res[res['prob']>0.01]
        print(res)
        crops = res.crop.values
        prob = res.prob.values
        def round_2(x):
            return round(x,2)
        prob = list(map(round_2, prob))
        final_res = zip(crops, prob)
        return render(request, 'crop_recom.html', {'final_res':final_res})

    return redirect('crop')

def login(request):
    if request.method == 'POST':
        email = request.POST['email']
        password = request.POST['password']

        user = auth.authenticate(username=email,password=password)
        if user is not None:
            auth.login(request,user)
            if 'next' in request.POST:
                return redirect(request.POST.get('next'))
            return redirect('index')
        else:
            messages.info(request,"Login Username or Password is wrong")
            return redirect("login")

    return render(request,"signup.html")

def signup(request):
    if request.method=='POST':
        email=request.POST['email']
        password=request.POST['password1']

        if User.objects.filter(username=email).exists():
                messages.info(request,"Signup Username is already taken!")
                return redirect('signup')
        else:
            user=User.objects.create_user(username=email,email=email,password=password)
            user.is_active = True
            user.save();
            user = auth.authenticate(username=email,password=password)
            auth.login(request,user)
            if 'next' in request.POST:
                return redirect(request.POST.get('next'))
            return redirect('index')

    return render(request,"signup.html")


def amaranth(request):
    return render(request,'amaranth.html')

def about(request):
    return render(request , 'about.html')

def crop(request):
    return render(request , 'crop_recom.html')



def logout(request):
    auth.logout(request)
    return redirect("index")

def index(request):
    if request.user.is_authenticated:
        user = request.user.username
        return render(request,'index.html',{'username':user})

    return render(request,'index.html')
    

def root(request):
    return redirect('index')

# def root(request):
#     print("In")
#     # Open file
#     book = openpyxl.load_workbook('Templates/Disease_Treatment.xlsx')
#     sheet = book.active
#     array = []
#     for i in range(1,sheet.max_row+1):
#         print(i)
#         for j in range(1 , sheet.max_column+1):
#             cell = sheet.cell(row=i , column=j)
#             array.append(cell.value)
#     print(len(array))
#     for i in range(0,len(array),2):
#         obj = diseaseInfo(
#             diseaseName = array[i],
#             context = array[i+1]
#         )
        
#         obj.save();
#     return render(request,'index.html')

# def root(request):
#     print("In")
#     # Open file
#     with open('Templates/plantInfo-clean.csv') as file_obj:
        
#         # Skips the heading
#         # Using next() method
#         heading = next(file_obj)
        
#         # Create reader object by passing the file 
#         # object to reader method
#         reader_obj = csv.reader(file_obj)
        
#         # Iterate over each row in the csv file 
#         # using reader object
#         for row in reader_obj:
#             print(row)
#             obj = cropDetails(
#                 name=row[0],
#                 alternateName=row[1],
#                 sowInstructions=row[2],
#                 spaceInstructions=row[3],
#                 harvestInstructions=row[4],
#                 compatiblePlants=row[5],
#                 avoidInstructions=row[6],
#                 culinaryHints=row[7],
#                 culinaryPreservation = row[8],
#                 url=row[9]
#             )
            
#             obj.save();
#         return render(request,'index.html')

